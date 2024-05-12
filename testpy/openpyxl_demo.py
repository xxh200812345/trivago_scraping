import openpyxl

# Load the workbook
wb = openpyxl.load_workbook(r'trivago_scraping\res\searchlist.xlsx')

# Get the active worksheet
ws = wb.active

# Print the column titles
for cell in ws[1]:
    print(cell.value)

# Print the row data
for index, row in enumerate(ws.iter_rows(min_row=2)):
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    print(f"{index} {row_data}")
