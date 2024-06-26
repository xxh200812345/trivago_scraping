import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('your_excel_file.xlsx')

# Get the active worksheet
ws = wb.active

# Print the column titles
for cell in ws[1]:
    print(cell.value)

# Print the row data
for row in ws.iter_rows(min_row=2):
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    print(row_data)
