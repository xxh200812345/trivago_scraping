#!/usr/bin/python3

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By

from selenium.common.exceptions import WebDriverException
import traceback

from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService

import yaml
from bs4 import BeautifulSoup
import openpyxl as px
from urllib.parse import urlparse, urlencode, urljoin

import logging
import logging.config
import time
import datetime
import os
import tkinter as tk
from tkinter import filedialog
import configparser

os.chdir(os.path.abspath(os.path.dirname(__file__)))

firstresearch = True
row_num = 0
filename = datetime.datetime.now().strftime("%Y%m%d%H%M")

driver = None
logger = None


def get_trivago_url(
    relative_url="hotels-tokyo-japan",
    city_no="200-71462",
    search_checkin="",
    search_checkout="",
    search_star="",
):
    """
    获取正确的URL
    """
    base_url = "https://www.trivago.hk/en-HK/lm/"

    # 获取当前日期和时间
    current_date = datetime.datetime.now()
    # 明天
    tomorrow = current_date + datetime.timedelta(days=1)
    # 后天
    day_after_tomorrow = current_date + datetime.timedelta(days=2)

    if search_checkin == "":
        search_checkin = tomorrow
    if search_checkout == "":
        search_checkout = day_after_tomorrow

    search_str = []

    if search_star:
        # 3星:105-1318 4星:105-1320 5星:105-1322
        if search_star == "3":
            search_str.append("105-1318")
        elif search_star == "4":
            search_str.append("105-1320")
        elif search_star == "5":
            search_str.append("105-1322")
        else:
            logger.info(f"星级输入错误：{search_star}")

    # hotels-tokyo-japan
    search_str.append(city_no)
    # dr-20230719-20230720
    search_str.append(
        "dr-"
        + search_checkin.strftime("%Y%m%d")
        + "-"
        + search_checkout.strftime("%Y%m%d")
    )

    # 200-71462;dr-20230719-20230720
    params = {"search": ";".join(search_str)}
    # search=200-71462;dr-20230719-20230720
    query_string = urlencode(params)
    # https://www.trivago.hk/en-HK/lm/hotels-tokyo-japan?search=200-71462;dr-20230719-20230720
    absolute_url = urljoin(base_url, relative_url + "?" + query_string)

    logger.info(f"生成查询URL：{absolute_url}")
    return absolute_url


def get_city_no(url):
    parsed_url = urlparse(url)

    print(f"path: {parsed_url.path}")
    print(f"query: {parsed_url.query}")

    relative_url = parsed_url.path.split("/")[-1]
    params = parsed_url.query.replace("search=", "")
    params = params.split(";")
    city_no = [s for s in params if s.startswith("200-")]
    ret = {"relative_url": relative_url, "city_no": city_no[0]}
    return ret


def has_element(byX, value):
    """
    检查元素是否存在
    """
    try:
        driver.find_element(byX, value)
        return True
    except:
        return False


def is_bot_wait():
    """
    是否被判定为机器人,True 是机器人
    """
    element = driver.find_element(By.XPATH, "//div[@id='sec-overlay']")
    return element.is_displayed()


def bot_check_wait(sleeptime=1):
    """
    如果被判定为机器人输入，等待
    """

    stime = 0
    time.sleep(sleeptime)
    if is_bot_wait():
        logger.info("反爬虫中...")
    # $x("")
    while is_bot_wait():
        time.sleep(1)
        # 超过31s依旧还在读取
        if stime > 31:
            driver.refresh()
            logger.info("超过31s依旧还在读取,刷新页面")
            stime = 0
            time.sleep(6)
        stime += 1
    if stime > 0:
        logger.info("反爬虫结束")


def check_next_btn():
    try:
        driver.find_element(By.XPATH, "//button[@data-testid='next-result-page']")
        return True
    except:
        return False


def element_click(element):
    """
    自定义按下事件
    """
    element.click()
    bot_check_wait()


def click_next_btn():
    element_click(
        driver.find_element(By.XPATH, "//button[@data-testid='next-result-page']")
    )
    WebDriverWait(driver, 15).until(EC.presence_of_all_elements_located)
    time.sleep(6)


def inittilt(nowsheet):
    nowsheet["A1"] = "Page-No"
    nowsheet.column_dimensions["A"].width = 15
    nowsheet["B1"] = "City"
    nowsheet.column_dimensions["B"].width = 15
    nowsheet["C1"] = "Hotelname"
    nowsheet.column_dimensions["C"].width = 50

    nowsheet["D1"] = "Recommend booking"
    nowsheet.column_dimensions["D"].width = 15
    nowsheet["E1"] = "price"
    nowsheet.column_dimensions["E"].width = 10
    nowsheet["F1"] = "Star"
    nowsheet.column_dimensions["F"].width = 10

    nowsheet["G1"] = "Other1"
    nowsheet.column_dimensions["G"].width = 15
    nowsheet["H1"] = "Other1"
    nowsheet.column_dimensions["H"].width = 10

    nowsheet["I1"] = "lowest booking"
    nowsheet.column_dimensions["I"].width = 15
    nowsheet["J1"] = "lowest price"
    nowsheet.column_dimensions["J"].width = 10

    nowsheet["K1"] = "Other3"
    nowsheet.column_dimensions["K"].width = 15
    nowsheet["L1"] = "Other3"
    nowsheet.column_dimensions["L"].width = 10

    nowsheet["M1"] = "Checkin"
    nowsheet.column_dimensions["M"].width = 15
    nowsheet["N1"] = "Checkout"
    nowsheet.column_dimensions["N"].width = 15

    nowsheet["O1"] = "Currency"
    nowsheet.column_dimensions["O"].width = 10


def scraping(
    search_cityname,
    search_checkin,
    search_checkout,
    real_currency,
    search_star,
    accommodation_list,
):
    """
    读取当前页的住宿清单
    """
    global nowsheet

    # 将滚动条移动到页面的最底部
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(2)
    driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

    i = nowsheet_i

    for page in accommodation_list:
        for item in page:
            if item:
                i += 1

    s = BeautifulSoup(driver.page_source, "html.parser")
    hotelitems = s.find_all("li", attrs={"data-testid": "accommodation-list-element"})

    n = 1
    page = []
    accommodation_list.append(page)

    for hotelitem in hotelitems:
        out_put = {}
        page.append(out_put)

        """
        星级信息筛选
        """
        stardiv = hotelitem.find("span", attrs={"itemprop": "starRating"})
        if stardiv != None:
            stardiv = stardiv.find("meta")["content"]
        else:
            stardiv = "0"
        if stardiv not in search_star:
            continue

        i += 1

        out_put["A" + str(i)] = (
            str(row_num - 1) + "-" + str(len(accommodation_list)) + "-" + str(n)
        )

        out_put["B" + str(i)] = search_cityname

        hotelname = hotelitem.find("span", attrs={"itemprop": "name"}).text
        out_put["C" + str(i)] = hotelname

        clickout_area = hotelitem.find("div", attrs={"data-testid": "clickout-area"})
        bookingsite = clickout_area.find("strong", attrs={"itemprop": "offeredBy"}).text
        out_put["D" + str(i)] = bookingsite

        hotelprice = clickout_area.find("p", attrs={"itemprop": "price"}).text
        hotelprice
        try:
            out_put["E" + str(i)] = int(hotelprice.lstrip("$").replace(",", ""))
        except:
            out_put["E" + str(i)] = hotelprice.lstrip("$").replace(",", "")

        star = hotelitem.find("meta")["content"]
        out_put["F" + str(i)] = star

        alternative_deal = hotelitem.find(
            "button", attrs={"data-testid": "alternative-deal"}
        )
        if alternative_deal != None:
            bookingsite = alternative_deal.find("span", attrs={"itemprop": "name"}).text
            out_put["G" + str(i)] = bookingsite

            hotelprice = alternative_deal.find("span", attrs={"class", "text-l"}).text
            try:
                out_put["H" + str(i)] = int(hotelprice.lstrip("$").replace(",", ""))
            except:
                out_put["H" + str(i)] = hotelprice.lstrip("$").replace(",", "")

        more_deals = hotelitem.find("button", attrs={"data-testid": "more-deals"})
        if more_deals != None:
            bookingsite = more_deals.find("span", attrs={"itemprop": "name"}).text
            out_put["I" + str(i)] = bookingsite

            hotelprice = more_deals.find("span", attrs={"class", "text-l"}).text
            try:
                out_put["J" + str(i)] = int(hotelprice.lstrip("$").replace(",", ""))
            except:
                out_put["J" + str(i)] = hotelprice.lstrip("$").replace(",", "")

        out_put["M" + str(i)] = search_checkin.strftime("%Y-%m-%d")
        out_put["N" + str(i)] = search_checkout.strftime("%Y-%m-%d")

        out_put["O" + str(i)] = real_currency

        n = n + 1


def inputcityname(cityname):
    """
    输入城市名
    """
    try:
        element_click(
            driver.find_element(
                By.XPATH, "//button[@data-testid='search-input-clear-button']"
            )
        )
        time.sleep(1)
    except:
        logger.info("不存在城市名删除按钮")

    driver.find_element(By.ID, "input-auto-complete").send_keys(cityname)
    bot_check_wait()
    driver.find_element(By.ID, "input-auto-complete").send_keys(Keys.ENTER)
    bot_check_wait()


def inputcurrency(search_currency):
    """
    变更货币类型
    """
    driver.delete_all_cookies()
    # Due to the internet speed, the currency selection will be loaded incorrectly. If the currency selection is wrong, please increase the sleep time here.
    # For example, change to time.sleep(2)
    time.sleep(1)

    # 输入货币类型
    show_currency = driver.find_element(
        By.XPATH, "//button[@data-testid='header-menu-currency-selector']/span[1]/span"
    ).text

    show_currency_btn = WebDriverWait(driver, timeout=30).until(
        lambda d: (d.find_element(By.XPATH, "//nav[1]/ul/li[4]"))
    )
    element_click(show_currency_btn)

    logger.info(f"当前显示货币类型：{show_currency}")

    if search_currency == show_currency:
        return

    select_element = WebDriverWait(driver, timeout=30).until(
        lambda d: (d.find_element(By.XPATH, "//select[@id='currency-selector']"))
    )
    select_obj = Select(select_element)
    select_obj.select_by_value(search_currency)
    bot_check_wait(3)

    logger.info(f"修正货币类型：{search_currency}")


def logging_init():
    global logger
    with open(file="logging.yml", mode="r", encoding="utf-8") as file:
        logging_yaml = yaml.load(stream=file, Loader=yaml.FullLoader)
        # print(logging_yaml)
        # 配置logging日志：主要从文件中读取handler的配置、formatter（格式化日志样式）、logger记录器的配置
        logging.config.dictConfig(config=logging_yaml)
    # 获取根记录器：配置信息从yaml文件中获取
    logger = logging.getLogger()
    # 子记录器的名字与配置文件中loggers字段内的保持一致
    my_module = logging.getLogger("my_module")
    print("rootlogger:", logger.handlers)
    print("selflogger", my_module.handlers)
    # print("子记录器与根记录器的handler是否相同：", root.handlers[0] == my_module.handlers[0])
    my_module.error("DUBUG")
    logger.info("" * 3)
    logger.info("-----------START")


def data_output_to_excel(accommodation_list):
    """
    住宿清单写入EXCEL
    """
    global nowsheet_i
    global nowsheet

    for i, page in enumerate(accommodation_list):
        # 初始化计数器
        count = 0
        for hotelitem in page:
            if hotelitem:
                count += 1
                for key, value in hotelitem.items():
                    nowsheet[key] = value
        nowsheet_i += count

        logger.info(f"## 第{(i+1)}页读取到{len(page)}个房源，实际输出到文件{count}个")


def body(search_sheet, nowsheet):
    """
    读取数据
    """
    search_cityname = search_sheet["A" + str(row_num)].value
    if search_cityname == None:
        logger.info(f"## 调查任务共计{row_num-2}件")
        return False

    search_cityname = search_cityname.strip()

    logger.info(f"## 读取第{row_num-2}条数据")

    search_checkin = search_sheet["B" + str(row_num)].value
    search_checkout = search_sheet["C" + str(row_num)].value
    # search_roomtype = search_sheet["D" + str(row_num)].value
    search_currency = search_sheet["E" + str(row_num)].value

    if search_sheet["F" + str(row_num)].value == None:
        search_star = ["3", "4", "5"]
    elif isinstance(search_sheet["F" + str(row_num)].value, int):
        search_star = str(search_sheet["F" + str(row_num)].value)
    else:
        search_star = search_sheet["F" + str(row_num)].value.split(",")

    logger.info(
        f"## 查询任务{row_num-2}({search_cityname,search_checkin.strftime('%Y-%m-%d'),search_checkout.strftime('%Y-%m-%d'),search_star})"
    )

    inittilt(nowsheet)

    inputcurrency(search_currency)
    time.sleep(3)
    driver.delete_all_cookies()

    # 城市信息
    inputcityname(search_cityname)
    city_no = get_city_no(driver.current_url)

    logger.info(f"## 输入检索城市信息：{search_cityname}，提取信息:{city_no}")

    url = get_trivago_url(
        city_no["relative_url"],
        city_no["city_no"],
        search_checkin,
        search_checkout,
        search_star,
    )

    driver.get(url)
    bot_check_wait(3)
    driver.get(url)

    # 住宿清单
    accommodation_list = []

    # 第一页
    scraping(
        search_cityname,
        search_checkin,
        search_checkout,
        search_currency,
        search_star,
        accommodation_list,
    )

    # 每一页
    while check_next_btn():
        click_next_btn()

        scraping(
            search_cityname,
            search_checkin,
            search_checkout,
            search_currency,
            search_star,
            accommodation_list,
        )

    # 住宿清单写入EXCEL
    data_output_to_excel(accommodation_list)

    return True


def driver_init_chrome():
    global driver

    op = webdriver.ChromeOptions()
    op.add_experimental_option("excludeSwitches", ["enable-automation"])
    op.add_experimental_option("useAutomationExtension", False)
    op.add_argument("--accept-language=en")

    driver = webdriver.Chrome(
        service=ChromeService(ChromeDriverManager().install()), options=op
    )
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {
            "source": """
        Object.defineProperty(navigator, 'webdriver', {
        get: () => undefined
        })
    """
        },
    )
    driver.execute_script("window.open('','_blank');")
    driver.switch_to.window(driver.window_handles[0])


def homepage_init():
    driver.get(get_trivago_url())
    driver.maximize_window()
    driver.implicitly_wait(2)
    driver.delete_all_cookies()

    bot_check_wait(3)


def config_init():
    global config
    # 创建一个ConfigParser对象
    config = configparser.ConfigParser()
    config.read(r"config.ini", encoding="utf-8")
    logging.info("读取配置文件成功。")


def open_file_dialog():
    root = tk.Tk()
    root.withdraw()

    file_path = config.get("Section", "file_path")

    # 文件存在
    if os.path.exists(file_path):
        logging.info(f"查询条件文件的路径是：{file_path}")
        return file_path

    # 弹出文件选择框
    file_path = filedialog.askopenfilename(
        title="Select Excel File for Query Conditions",
        filetypes=[("Excel文件", "*.xlsx;*.xls")],
    )

    # 更新配置文件中的路径
    config.set("Section", "file_path", file_path)

    # 写入配置文件
    with open("config.ini", "w") as configfile:
        config.write(configfile)

        logging.info(f"文件路径 {file_path} 已保存到 config.ini文件中。")

    return file_path


if __name__ == "__main__":  # ⼊⼝函数
    logging_init()

    config_init()
    file_path = open_file_dialog()

    logger.info(config.get("Section", "version"))

    # 创建文件夹
    result_dir = "output"
    if not os.path.exists(result_dir):
        os.makedirs(result_dir)
        logger.info(f"创建{result_dir}")

    driver_init_chrome()
    homepage_init()

    # 取searchlist数据
    searchlist_exl = px.load_workbook(file_path)
    search_sheet = searchlist_exl.active

    # 保存结果的excel生成
    retexl = px.Workbook()
    nowsheet = retexl.active
    nowsheet_i = 1

    logger.info(f"# 调查任务开始")

    row_num = 2
    while row_num < search_sheet.max_row + 1:
        try:
            if body(search_sheet, nowsheet) is False:
                break
            retexl.save(f"{result_dir}/{filename}.xlsx")
            row_num += 1
        except WebDriverException as e:
            logger.info(e)
            logger.info("出现WebDriverException错误，重新打开新的浏览器.")
            traceback.print_exc()
            if driver is not None:
                driver.quit()
            driver_init_chrome()
            homepage_init()

    retexl.save(f"{result_dir}/{filename}.xlsx")

    driver.quit()
    logger.info("Finished! Press Enter to close...")
