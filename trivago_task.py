import time
import random
from typing import List
from datetime import datetime

from trivago_log import TaLog
from trivago_tool import TaConfig

import openpyxl as px
from openpyxl import Workbook
from openpyxl.styles import NamedStyle


class TaTask:
    cityname: str = None
    checkin: datetime = None
    checkout: datetime = None
    roomtype: str = None
    currency: str = None
    star: int = None
    location_slug: str = None
    city_code: str = None
    log_key = None
    line_no = None
    url = None
    ROOM_TYPE_SINGLE = "Single room"
    state = None
    search_start_time = None

    STATE_NORMAL = "等待处理"
    STATE_ERROR = "数据错误"
    STATE_OVER = "处理结束"

    # init
    def __init__(self, cell: list, index: int):
        try:
            self.line_no = f"line[{(index + 2):06}]"
            fields = ["cityname", "checkin", "checkout", "roomtype", "currency", "star", "location_slug", "city_code"]
            for i, field in enumerate(fields):
                setattr(self, field, cell[i] if i < len(cell) else "")
            self.roomtype = TaTask.ROOM_TYPE_SINGLE
            self.state = TaTask.STATE_NORMAL

        except Exception as e:
            TaLog().error(f"{self.log_key}TaTask init error: {e}")
            TaLog().error(f"{self.log_key}{cell}")

            self.state = TaTask.STATE_ERROR

    # print
    def __repr__(self):
        return str(self.__dict__)

    @staticmethod
    def get_tasks(file_path):
        # 取searchlist数据
        searchlist_exl = px.load_workbook(file_path)
        ws = searchlist_exl.active

        tasks = []
        # Print the row data
        for index, row in enumerate(ws.iter_rows(min_row=2)):
            row_data = []
            if all(cell.value is None for cell in row):
                continue
            for cell in row:
                row_data.append(cell.value)
            _task = TaTask(row_data, index)
            if _task.cityname != None:
                tasks.append(_task)

        # 2. 动态计算补零宽度
        total_count = len(tasks)
        # 计算总数的位数，比如 500 是 3 位，1000 是 4 位
        width = len(str(total_count)) 

        # 3. 统一格式化 log_key
        for i, task in enumerate(tasks):
            current_num = i + 1
            task.log_key = f"[{current_num:0{width}d}/{total_count}] "
        return tasks

    def check_roomtype(self, roomtype: str):
        if roomtype != self.ROOM_TYPE_SINGLE:
            raise ValueError("房间类型错误")
        return roomtype
    

    @staticmethod
    def output_create():
        config = TaConfig().config
        filename = datetime.now().strftime("%Y%m%d%H%M")
        output_dir = config["output"]["path"]
        TaConfig.ensure_file_exists(output_dir)

        output_file = f"{output_dir}{filename}.xlsx"
        wb = Workbook()
        # 获取活动工作表
        ws = wb.active

        # 设置工作表标题
        ws.title = "Sheet1"

        if "date_style" not in wb.named_styles:
            date_style = NamedStyle(name="date_style", number_format="YYYY年MM月DD日")
            wb.add_named_style(date_style)

        # 将标题写入第一行
        titles = config["output"]["titles"]
        for col_num, title in enumerate(titles, start=1):
            ws.cell(row=1, column=col_num, value=title)
        wb.save(output_file)
        return output_file

    @staticmethod
    def output(file_path, outputs: list[dict]):
        # 取searchlist数据
        wb = px.load_workbook(file_path)
        ws = wb.active

        # 找到最后一行
        last_row = ws.max_row + 1

        config = TaConfig().config
        titles = config["output"]["titles"]
        titles_dict = {title: index + 1 for index, title in enumerate(titles)}

        for output in outputs:
            # 将新数据写入最后一行
            for _, title in enumerate(output, start=1):
                cell = ws.cell(
                    row=last_row,
                    column=titles_dict.get(title),
                    value=output.get(title, ""),
                )
                if type(cell.value) == datetime:
                    cell.style = "date_style"

            last_row += 1

        # 保存工作簿
        wb.save(file_path)

    def checkin_for_url(self):
        return self.checkin.strftime("%Y%m%d")

    def checkout_for_url(self):
        return self.checkout.strftime("%Y%m%d")

    def star_for_url(self):
        config = TaConfig().config
        stars = config["stars"]
        return stars[self.star]
    
    def record_start_time(self):
        """
        将当前系统时间记录到 search_start_time 属性中
        """
        self.search_start_time = datetime.now()
        TaLog().info(f"{self.log_key}开始执行任务时间已记录: {self.search_start_time}")

    def wait_until_interval(self, interval_seconds: int = 10):
        """
        判定当前时间与 search_start_time 的差值。
        如果不足指定秒数（含上下10%浮动），则进行等待。
        """
        if self.search_start_time is None:
            return

        # 加入随机浮动：例如 10s 会变成 9s ~ 11s 之间的随机数
        random_factor = random.uniform(0.9, 1.1)
        actual_interval = interval_seconds * random_factor

        # 计算自上次记录 search_start_time 以来已流逝的时间
        elapsed_time = (datetime.now() - self.search_start_time).total_seconds()
        
        if elapsed_time < actual_interval:
            wait_time = actual_interval - elapsed_time
            # 这里的日志可以打印出目标随机间隔，方便调试
            TaLog().info(f"{self.log_key} 随机目标间隔为 {actual_interval:.2f}s (已过去 {elapsed_time:.2f}s)，需补足等待: {wait_time:.2f}s")
            time.sleep(wait_time)
        else:
            TaLog().info(f"{self.log_key} 间隔已达标 (目标 {actual_interval:.2f}s，已过去 {elapsed_time:.2f}s)，无需额外等待")

    @staticmethod
    def count_task_states(tasks: List["TaTask"]) -> dict:
        state_counts = {
            TaTask.STATE_NORMAL: 0,
            TaTask.STATE_ERROR: 0,
            TaTask.STATE_OVER: 0,
        }

        for task in tasks:
            if task.state == TaTask.STATE_NORMAL:
                state_counts[TaTask.STATE_NORMAL] += 1
            elif task.state == TaTask.STATE_ERROR:
                state_counts[TaTask.STATE_ERROR] += 1
            elif task.state == TaTask.STATE_OVER:
                state_counts[TaTask.STATE_OVER] += 1

        TaLog().info(state_counts)


def test():
    # Exam  ple usage:
    task1 = TaTask()
    task1.state = TaTask.STATE_NORMAL

    task2 = TaTask()
    task2.state = TaTask.STATE_ERROR

    task3 = TaTask()
    task3.state = TaTask.STATE_OVER

    task4 = TaTask()
    task4.state = TaTask.STATE_NORMAL

    task_list = [task1, task2, task3, task4]
    TaTask.count_task_states(task_list)
